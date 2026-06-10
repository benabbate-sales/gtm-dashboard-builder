#!/usr/bin/env python3
"""Build a three-dashboard GTM workbook (Quarterly Forecast, FY Forecast &
Pipeline, Seller Performance) from flat CRM export CSVs + a client config.

Usage:
  python3 build_dashboards.py --config config.json \
      --opportunities opportunities.csv [--line-items line_items.csv] \
      [--targets targets.csv] [--as-of YYYY-MM-DD] --output out.xlsx

Requires: pandas, openpyxl. See references/dashboard-specs.md for the
metric definitions this implements.
"""
import argparse
import datetime as dt
import json
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling
DARK = "1F3864"      # section title band
MID = "D6E4F0"       # column header fill
AMBER = "FFE699"     # missing-input highlight
GREY = "F2F2F2"      # note bars
THIN = Side(style="thin", color="BFBFBF")
TOP = Border(top=Side(style="medium", color="1F3864"))

F_TITLE = Font(bold=True, color="FFFFFF", size=12)
F_HDR = Font(bold=True, size=10)
F_NOTE = Font(italic=True, size=9, color="595959")
F_BOLD = Font(bold=True)


def section_title(ws, row, text, width):
    ws.cell(row=row, column=1, value=text).font = F_TITLE
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=DARK)
    return row + 1


def note(ws, row, text, width):
    ws.cell(row=row, column=1, value=text).font = F_NOTE
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=GREY)
    return row + 1


def write_df(ws, row, df, fmts=None, total_row=False):
    """Write a dataframe with styled headers. fmts maps column name -> number
    format. If total_row, the last row is bolded with a top border."""
    fmts = fmts or {}
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=row, column=j, value=col)
        c.font = F_HDR
        c.fill = PatternFill("solid", fgColor=MID)
        c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r = row + 1
    n = len(df)
    for i, (_, rec) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns, start=1):
            v = rec[col]
            if pd.isna(v):
                v = None
            c = ws.cell(row=r, column=j, value=v)
            if col in fmts and isinstance(v, (int, float)):
                c.number_format = fmts[col]
            if total_row and i == n - 1:
                c.font = F_BOLD
                c.border = TOP
        r += 1
    return r + 1


def autosize(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value is not None),
                      default=0)
        ws.column_dimensions[letter].width = min(max(10, longest + 2), max_width)


# ---------------------------------------------------------------- fiscal
def fiscal_quarter(d, start_month):
    """Return (fy, q) for a date. FY labelled by the calendar year it starts."""
    offset = (d.month - start_month) % 12
    fy = d.year if d.month >= start_month else d.year - 1
    if start_month == 1:
        fy = d.year
    return fy, offset // 3 + 1


def q_label(fy, q):
    return f"FY{fy} Q{q}"


def quarter_months(fy, q, start_month):
    m0 = (start_month - 1 + (q - 1) * 3) % 12
    months = []
    for k in range(3):
        m = (m0 + k) % 12 + 1
        y = fy + ((m0 + k) // 12)
        if m < start_month:
            y = fy + 1
        if start_month == 1:
            y = fy
        months.append((y, m))
    return months


def prev_quarter(fy, q, n=1):
    for _ in range(n):
        q -= 1
        if q == 0:
            fy, q = fy - 1, 4
    return fy, q


# ---------------------------------------------------------------- load
def load(args):
    cfg = json.load(open(args.config))
    df = pd.read_csv(args.opportunities)
    df.columns = [c.strip() for c in df.columns]
    required = ["Opp Id", "Opportunity Name", "Opp Owner", "Stage",
                "Forecast Category", "Value", "Close Date"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        sys.exit(f"opportunities.csv missing required columns: {missing}")

    df["Value"] = pd.to_numeric(df["Value"], errors="coerce").fillna(0.0)
    df["Close Date"] = pd.to_datetime(df["Close Date"], errors="coerce")
    if "Created Date" in df.columns:
        df["Created Date"] = pd.to_datetime(df["Created Date"], errors="coerce")

    open_stages = [s["name"] for s in cfg["stages"]]
    probs = {s["name"]: s["probability"] for s in cfg["stages"]}
    probs[cfg["won_stage"]] = 1.0
    probs[cfg["lost_stage"]] = 0.0
    known = set(probs)
    unknown = sorted(set(df["Stage"].dropna()) - known)

    df["__prob"] = df["Stage"].map(probs).fillna(0.0)
    df["Weighted"] = df["Value"] * df["__prob"]
    df["__open"] = df["Stage"].isin(open_stages)
    df["__won"] = df["Stage"] == cfg["won_stage"]
    df["__lost"] = df["Stage"] == cfg["lost_stage"]

    fc = cfg["forecast_categories"]
    rev = {v: k for k, v in fc.items()}
    df["__bucket"] = df["Forecast Category"].map(rev)

    inc = cfg.get("use_include_forecast_flag") and "Include Forecast" in df.columns
    if inc:
        df["__inc"] = pd.to_numeric(df["Include Forecast"], errors="coerce").fillna(0) == 1
    else:
        df["__inc"] = df["__bucket"] == "best_case"  # all BC counts as strong

    bands = sorted(cfg.get("deal_bands", []), key=lambda b: b["min"])
    if bands:
        def band(v):
            name = bands[0]["name"]
            for b in bands:
                if v >= b["min"]:
                    name = b["name"]
            return name
        df["Deal Band"] = df["Value"].apply(band)

    sm = cfg.get("fiscal_year_start_month", 1)
    fq = df["Close Date"].apply(
        lambda d: fiscal_quarter(d, sm) if pd.notna(d) else (None, None))
    df["__fy"] = [t[0] for t in fq]
    df["__q"] = [t[1] for t in fq]

    li = None
    if args.line_items:
        li = pd.read_csv(args.line_items)
        li.columns = [c.strip() for c in li.columns]
        li["Product Value"] = pd.to_numeric(li.get("Product Value"),
                                            errors="coerce").fillna(0.0)

    tg = None
    if args.targets:
        tg = pd.read_csv(args.targets)
        tg.columns = [c.strip() for c in tg.columns]
        tg["Owner"] = tg.get("Owner", pd.Series(dtype=str)).fillna("")
        tg["Target"] = pd.to_numeric(tg["Target"], errors="coerce")

    return cfg, df, li, tg, unknown


def get_target(tg, period, owner=""):
    if tg is None:
        return None
    m = tg[(tg["Period"].astype(str).str.strip() == period) &
           (tg["Owner"].astype(str).str.strip() == owner)]
    if m.empty:
        return None
    return float(m["Target"].iloc[0])


# ---------------------------------------------------------------- rollups
def confidence_row(d, label):
    won = d[d["__won"]]
    commit = d[d["__open"] & (d["__bucket"] == "commit")]
    bc = d[d["__open"] & (d["__bucket"] == "best_case")]
    strong = bc[bc["__inc"]]
    pipe = d[d["__open"] & (d["__bucket"] == "pipeline")]
    total = won["Value"].sum() + commit["Value"].sum() + bc["Value"].sum() + pipe["Value"].sum()
    return {
        "Period": label,
        "Won": won["Value"].sum(), "# Won": len(won),
        "Commit": commit["Value"].sum(), "# Commit": len(commit),
        "Strong Best Case": strong["Value"].sum(),
        "Remaining Best Case": bc["Value"].sum() - strong["Value"].sum(),
        "# Best Case": len(bc),
        "Pipeline": pipe["Value"].sum(), "# Pipeline": len(pipe),
        "Total Pipeline": total,
        "Weighted Pipeline": d[d["__open"]]["Weighted"].sum() + won["Value"].sum(),
    }


def add_totals(rows, label="Total"):
    if not rows:
        return rows
    tot = {}
    for k in rows[0]:
        vals = [r[k] for r in rows]
        tot[k] = sum(vals) if all(isinstance(v, (int, float)) for v in vals) else label
    rows.append(tot)
    return rows


CUR_COLS = ["Won", "Commit", "Strong Best Case", "Remaining Best Case",
            "Pipeline", "Total Pipeline", "Weighted Pipeline", "Value",
            "Weighted", "Bookings", "Quota", "Forecast", "ASP", "Target"]


def fmts_for(df_cols, cur):
    f = {}
    for c in df_cols:
        if any(c.startswith(p) or c.endswith(p) for p in CUR_COLS) or "Value" in c or "ARR" in c:
            f[c] = f'"{cur}"#,##0'
        if c.startswith("#") or c.endswith("# Opps"):
            f[c] = "#,##0"
        if "%" in c or "Attainment" in c:
            f[c] = "0%"
        if "Coverage" in c:
            f[c] = '0.0"x"'
    return f


# ---------------------------------------------------------------- KPI grid
def kpi_grid(ws, row, kpis, cur, width):
    """kpis: list of (label, value, sub) tuples; 4 per row."""
    per = 4
    span = 3
    for i in range(0, len(kpis), per):
        chunk = kpis[i:i + per]
        for j, (label, value, sub) in enumerate(chunk):
            c0 = j * span + 1
            lc = ws.cell(row=row, column=c0, value=label)
            lc.font = F_HDR
            vc = ws.cell(row=row + 1, column=c0, value=value)
            vc.font = Font(bold=True, size=14)
            if isinstance(value, (int, float)):
                vc.number_format = f'"{cur}"#,##0'
            else:
                vc.fill = PatternFill("solid", fgColor=AMBER)
            sc = ws.cell(row=row + 2, column=c0, value=sub)
            sc.font = F_NOTE
        row += 4
    return row + 1


def fmt_pct_of(value, target):
    if target in (None, 0):
        return "awaiting target"
    return f"{value / target:.0%} of target · gap {target - value:,.0f}"


# ---------------------------------------------------------------- tabs
def tab_quarterly(wb, cfg, df, tg, asof):
    sm = cfg.get("fiscal_year_start_month", 1)
    fy, q = fiscal_quarter(asof, sm)
    ql = q_label(fy, q)
    cur = cfg.get("currency_symbol", "$")
    vl = cfg.get("value_label", "Value")
    ws = wb.create_sheet("1. Quarterly Forecast")
    W = 13

    row = section_title(ws, 1, f"QUARTERLY SALES FORECAST — {ql}", W)
    row = note(ws, row, f"As of {asof:%d %b %Y}. Value measure: {vl}. "
                        f"Scope: opportunities with Close Date in {ql}.", W)
    row += 1

    dq = df[(df["__fy"] == fy) & (df["__q"] == q)]
    r0 = confidence_row(dq, ql)
    target = get_target(tg, ql)

    open_q = dq[dq["__open"]]
    won_v = r0["Won"]
    remaining = None if target is None else max(target - won_v, 0)
    wcov = (open_q["Weighted"].sum() / remaining) if remaining else None
    strong_total = r0["Strong Best Case"]
    scov = (strong_total / remaining) if remaining else None

    kpis = [
        ("Won " + vl, won_v, fmt_pct_of(won_v, target)),
        ("Commit + Won", won_v + r0["Commit"], fmt_pct_of(won_v + r0["Commit"], target)),
        ("Strong BC + Commit + Won", won_v + r0["Commit"] + strong_total,
         fmt_pct_of(won_v + r0["Commit"] + strong_total, target)),
        ("Weighted Pipeline", r0["Weighted Pipeline"], fmt_pct_of(r0["Weighted Pipeline"], target)),
        ("Pipe Coverage (weighted)",
         "awaiting target" if wcov is None else round(wcov, 2),
         "weighted open pipe ÷ remaining-to-target"),
        ("Pipe Coverage (strong best case)",
         "awaiting target" if scov is None else round(scov, 2),
         "strong best case ÷ remaining-to-target"),
    ]
    for flag in cfg.get("deal_quality_flags", []):
        col, lbl = flag["column"], flag["label"]
        if col not in df.columns:
            continue
        min_stage = cfg["stages"][flag.get("min_stage_index", 0)]["name"]
        idx = [s["name"] for s in cfg["stages"]].index(min_stage)
        elig_stages = [s["name"] for s in cfg["stages"][idx:]]
        elig = dq[dq["Stage"].isin(elig_stages) &
                  (dq["Value"] >= flag.get("min_value", 0))]
        done = elig[pd.to_numeric(elig[col], errors="coerce").fillna(0) == 1]
        kpis.append((f"{lbl} — done / eligible", f"{len(done)} / {len(elig)}",
                     f"stage ≥ {min_stage}, {vl} ≥ {cur}{flag.get('min_value', 0):,.0f}"))
    row = kpi_grid(ws, row, kpis, cur, W)

    row = section_title(ws, row, "Forecast Confidence by Month", W)
    months = quarter_months(fy, q, sm)
    rows = []
    for (y, m) in months:
        dm = dq[(dq["Close Date"].dt.year == y) & (dq["Close Date"].dt.month == m)]
        rows.append(confidence_row(dm, dt.date(y, m, 1).strftime("%b %Y")))
    add_totals(rows)
    t = pd.DataFrame(rows)
    row = write_df(ws, row, t, fmts_for(t.columns, cur), total_row=True)

    row = section_title(ws, row, "Weighted Forecast — Stage × Month", W)
    stage_cols = [cfg["won_stage"]] + [s["name"] for s in cfg["stages"]]
    rows = []
    for (y, m) in months:
        dm = dq[(dq["Close Date"].dt.year == y) & (dq["Close Date"].dt.month == m)]
        rec = {"Month": dt.date(y, m, 1).strftime("%b %Y")}
        for s in stage_cols:
            rec[s] = dm[dm["Stage"] == s]["Weighted"].sum()
        rec["Grand total"] = sum(rec[s] for s in stage_cols)
        rows.append(rec)
    add_totals(rows)
    t = pd.DataFrame(rows)
    row = write_df(ws, row, t, {c: f'"{cur}"#,##0' for c in t.columns[1:]}, total_row=True)

    row = section_title(ws, row, "Pipeline by Stage × Forecast Category", W)
    rows = []
    for bk, label in cfg["forecast_categories"].items():
        db = dq[dq["__bucket"] == bk]
        rec = {"Forecast Category": label}
        for s in stage_cols:
            rec[s] = db[db["Stage"] == s]["Value"].sum()
        rec["Grand total"] = sum(rec[s] for s in stage_cols)
        rows.append(rec)
    add_totals(rows)
    t = pd.DataFrame(rows)
    row = write_df(ws, row, t, {c: f'"{cur}"#,##0' for c in t.columns[1:]}, total_row=True)

    row = section_title(ws, row, "Closed Won — this quarter", W)
    won = dq[dq["__won"]].sort_values("Value", ascending=False)
    cols = [c for c in ["Opp Owner", "Type", "Opportunity Name", "Account Name",
                        "ICP Grade", "Close Date", "Value", "Opp Link"] if c in won.columns]
    t = won[cols].copy()
    if "Created Date" in won.columns:
        t.insert(len(cols) - 1, "Deal Cycle (days)",
                 (won["Close Date"] - won["Created Date"]).dt.days)
    t["Close Date"] = t["Close Date"].dt.strftime("%Y-%m-%d")
    row = write_df(ws, row, t, fmts_for(t.columns, cur))

    row = section_title(ws, row, "Moved to Lost — last 7 days", W)
    lost = df[df["__lost"] & (df["Close Date"] >= pd.Timestamp(asof) - pd.Timedelta(days=7))
              & (df["Close Date"] <= pd.Timestamp(asof))]
    cols = [c for c in ["Opp Owner", "Type", "Opportunity Name", "ICP Grade",
                        "Lead Source", "Loss Reason", "Close Date", "Value"] if c in lost.columns]
    t = lost[cols].copy()
    if len(t):
        t["Close Date"] = pd.to_datetime(t["Close Date"]).dt.strftime("%Y-%m-%d")
    row = write_df(ws, row, t, fmts_for(t.columns, cur))

    for flag in cfg.get("deal_quality_flags", []):
        col = flag["column"]
        if col not in df.columns:
            continue
        idx = flag.get("min_stage_index", 0)
        elig_stages = [s["name"] for s in cfg["stages"][idx:]]
        miss = df[df["__open"] & df["Stage"].isin(elig_stages)
                  & (df["Value"] >= flag.get("min_value", 0))
                  & (pd.to_numeric(df[col], errors="coerce").fillna(0) != 1)]
        row = section_title(ws, row, f"Eligible opps missing {flag['label']}", W)
        cols = [c for c in ["Opp Owner", "Forecast Category", "Stage",
                            "Opportunity Name", "Next Step", "Close Date", "Value"]
                if c in miss.columns]
        t = miss[cols].sort_values("Value", ascending=False).copy()
        if len(t):
            t["Close Date"] = pd.to_datetime(t["Close Date"]).dt.strftime("%Y-%m-%d")
        row = write_df(ws, row, t, fmts_for(t.columns, cur))

    row = section_title(ws, row, "Deal Review — open opps closing this quarter", W)
    rev = dq[dq["__open"]].sort_values("Value", ascending=False)
    cols = [c for c in ["Opp Owner", "Forecast Category", "Stage", "Opportunity Name",
                        "Next Step", "Type", "Deal Band", "ICP Grade", "Close Date",
                        "Value", "Opp Link"] if c in rev.columns]
    t = rev[cols].copy()
    if len(t):
        t["Close Date"] = pd.to_datetime(t["Close Date"]).dt.strftime("%Y-%m-%d")
    write_df(ws, row, t, fmts_for(t.columns, cur))
    ws.freeze_panes = "A3"
    autosize(ws)


def tab_fy(wb, cfg, df, li, tg, asof):
    sm = cfg.get("fiscal_year_start_month", 1)
    fy = cfg.get("fiscal_year", fiscal_quarter(asof, sm)[0])
    fyl = cfg.get("fy_label", f"FY{fy}")
    cur = cfg.get("currency_symbol", "$")
    vl = cfg.get("value_label", "Value")
    ws = wb.create_sheet("2. FY Forecast & Pipeline")
    W = 13

    row = section_title(ws, 1, f"FY SALES FORECAST & PIPELINE — {fyl}", W)
    row = note(ws, row, f"As of {asof:%d %b %Y}. Value measure: {vl}.", W)
    row += 1

    dfy = df[df["__fy"] == fy]
    r0 = confidence_row(dfy, fyl)
    target = get_target(tg, f"FY{fy}")
    open_all = df[df["__open"]]

    kpis = [
        (f"{fyl} Bookings", r0["Won"], fmt_pct_of(r0["Won"], target)),
        ("Commit + Won", r0["Won"] + r0["Commit"],
         fmt_pct_of(r0["Won"] + r0["Commit"], target)),
        ("Strong BC + Commit + Won",
         r0["Won"] + r0["Commit"] + r0["Strong Best Case"],
         fmt_pct_of(r0["Won"] + r0["Commit"] + r0["Strong Best Case"], target)),
        ("Weighted Pipeline", r0["Weighted Pipeline"],
         fmt_pct_of(r0["Weighted Pipeline"], target)),
        ("Total Open Pipeline (all dates)", open_all["Value"].sum(),
         f"{len(open_all)} open opps, close date unbounded"),
        ("# Open Opps in FY", int(dfy["__open"].sum()), ""),
    ]
    row = kpi_grid(ws, row, kpis, cur, W)

    row = section_title(ws, row, "Forecast Confidence by Quarter", W)
    rows = []
    for qq in range(1, 5):
        dqq = dfy[dfy["__q"] == qq]
        rows.append(confidence_row(dqq, q_label(fy, qq)))
    add_totals(rows)
    t = pd.DataFrame(rows)
    row = write_df(ws, row, t, fmts_for(t.columns, cur), total_row=True)

    if "Deal Band" in df.columns:
        row = section_title(ws, row, "Open Pipeline by Deal Band × Quarter", W)
        bands = [b["name"] for b in sorted(cfg["deal_bands"], key=lambda b: b["min"])]
        rows = []
        for qq in range(1, 5):
            dqq = dfy[(dfy["__q"] == qq) & dfy["__open"]]
            rec = {"Quarter": q_label(fy, qq)}
            for b in bands:
                sel = dqq[dqq["Deal Band"] == b]
                rec[f"{b} {vl}"] = sel["Value"].sum()
                rec[f"{b} # Opps"] = len(sel)
            rec[f"Total {vl}"] = dqq["Value"].sum()
            rec["Total # Opps"] = len(dqq)
            rows.append(rec)
        add_totals(rows)
        t = pd.DataFrame(rows)
        row = write_df(ws, row, t, fmts_for(t.columns, cur), total_row=True)

    prod_col = None
    if li is not None and "Product" in li.columns:
        merged = li.merge(df[["Opp Id", "__fy", "__q", "__open"]], on="Opp Id", how="left")
        merged = merged[(merged["__fy"] == fy) & merged["__open"]]
        if len(merged):
            row = section_title(ws, row, "Open Pipeline by Product × Quarter", W)
            piv = merged.pivot_table(index="__q", columns="Product",
                                     values="Product Value", aggfunc="sum",
                                     fill_value=0)
            rows = []
            for qq in range(1, 5):
                rec = {"Quarter": q_label(fy, qq)}
                for p in piv.columns:
                    rec[p] = float(piv.loc[qq, p]) if qq in piv.index else 0.0
                rec["Grand total"] = sum(v for k, v in rec.items() if k != "Quarter")
                rows.append(rec)
            add_totals(rows)
            t = pd.DataFrame(rows)
            row = write_df(ws, row, t, {c: f'"{cur}"#,##0' for c in t.columns[1:]},
                           total_row=True)
    elif "Product" in df.columns:
        prod_col = "Product"
    if li is None and prod_col is None:
        row = note(ws, row, "By-product view skipped — no line-item or product "
                            "data supplied.", W)
        row += 1

    stage_cols = [cfg["won_stage"]] + [s["name"] for s in cfg["stages"]]
    for title, field in [("Weighted Forecast — Stage × Quarter", "Weighted"),
                         ("Total Pipeline — Stage × Quarter", "Value")]:
        row = section_title(ws, row, title, W)
        rows = []
        for qq in range(1, 5):
            dqq = dfy[dfy["__q"] == qq]
            rec = {"Quarter": q_label(fy, qq)}
            for s in stage_cols:
                rec[s] = dqq[dqq["Stage"] == s][field].sum()
            rec["Grand total"] = sum(rec[s] for s in stage_cols)
            rows.append(rec)
        add_totals(rows)
        t = pd.DataFrame(rows)
        row = write_df(ws, row, t, {c: f'"{cur}"#,##0' for c in t.columns[1:]},
                       total_row=True)

    if "Created Date" in df.columns:
        row = section_title(ws, row, "Pipeline Created by Month", W)
        created = df[df["Created Date"].notna()].copy()
        created["__cm"] = created["Created Date"].dt.to_period("M")
        fy_start = pd.Timestamp(dt.date(fy, sm, 1))
        created = created[created["Created Date"] >= fy_start]
        grp = created.groupby("__cm").agg(**{
            f"Created {vl}": ("Value", "sum"), "# Opps": ("Opp Id", "count")})
        grp = grp.reset_index()
        grp["__cm"] = grp["__cm"].astype(str)
        grp.columns = ["Created Month", f"Created {vl}", "# Opps"]
        row = write_df(ws, row, grp, fmts_for(grp.columns, cur))
    ws.freeze_panes = "A3"
    autosize(ws)


def tab_sellers(wb, cfg, df, tg, asof):
    sm = cfg.get("fiscal_year_start_month", 1)
    fy = cfg.get("fiscal_year", fiscal_quarter(asof, sm)[0])
    cfy, cq = fiscal_quarter(asof, sm)
    cql = q_label(cfy, cq)
    cur = cfg.get("currency_symbol", "$")
    vl = cfg.get("value_label", "Value")
    ws = wb.create_sheet("3. Seller Performance")
    W = 16

    row = section_title(ws, 1, "SELLER PERFORMANCE", W)
    row = note(ws, row, f"As of {asof:%d %b %Y}. Quota and targets are supplied "
                        f"inputs (targets.csv), never CRM data.", W)
    row += 1

    dfy = df[df["__fy"] == fy]
    owners = sorted(dfy["Opp Owner"].dropna().unique())

    row = section_title(ws, row, f"Forecast by Owner — FY{fy}", W)
    rows = []
    for o in owners:
        rec = confidence_row(dfy[dfy["Opp Owner"] == o], o)
        rec["Owner"] = rec.pop("Period")
        rows.append({"Owner": rec["Owner"], **{k: v for k, v in rec.items() if k != "Owner"}})
    add_totals(rows)
    t = pd.DataFrame(rows)
    row = write_df(ws, row, t, fmts_for(t.columns, cur), total_row=True)

    row = section_title(ws, row, f"Quarterly Scorecard — {cql}", W)
    row = note(ws, row, "Attainment = Bookings ÷ Quota. Forecast = Won + Commit. "
                        "ASP = Bookings ÷ # Won. Blank quota = awaiting target.", W)
    rows = []
    dq = df[(df["__fy"] == cfy) & (df["__q"] == cq)]
    large = cfg.get("large_bands", [])
    for o in owners:
        d = dq[dq["Opp Owner"] == o]
        won = d[d["__won"]]
        bookings = won["Value"].sum()
        quota = get_target(tg, cql, o)
        commit = d[d["__open"] & (d["__bucket"] == "commit")]["Value"].sum()
        fcst = bookings + commit
        wpipe = d[d["__open"]]["Weighted"].sum()
        remaining = None if quota in (None, 0) else max(quota - bookings, 0)
        open_d = d[d["__open"]]
        lg = open_d[open_d.get("Deal Band", pd.Series(dtype=str)).isin(large)] \
            if "Deal Band" in d.columns else open_d.iloc[0:0]
        rows.append({
            "Sales Rep": o,
            "Attainment": (bookings / quota) if quota else None,
            "Quota": quota, "Bookings": bookings,
            "ASP": (bookings / len(won)) if len(won) else None,
            "Forecast Attainment": (fcst / quota) if quota else None,
            "Forecast": fcst, "Weighted Pipeline": wpipe,
            "Weighted Pipe Coverage": (wpipe / remaining) if remaining else None,
            f"Large-deal {vl}": lg["Value"].sum(),
            "% Large-deal": (lg["Value"].sum() / open_d["Value"].sum())
            if open_d["Value"].sum() else None,
        })
    t = pd.DataFrame(rows)
    f = fmts_for(t.columns, cur)
    f.update({"Attainment": "0%", "Forecast Attainment": "0%", "% Large-deal": "0%",
              "Weighted Pipe Coverage": '0.0"x"'})
    row = write_df(ws, row, t, f)

    row = section_title(ws, row, "Trailing + Current Quarter Scorecard", W)
    p1 = prev_quarter(cfy, cq, 1)
    p2 = prev_quarter(cfy, cq, 2)
    rows = []
    for o in owners:
        trail = df[(df[["__fy", "__q"]].apply(tuple, axis=1).isin([p1, p2]))
                   & (df["Opp Owner"] == o)]
        twon = trail[trail["__won"]]
        tquota = sum(filter(None, [get_target(tg, q_label(*p), o) for p in (p1, p2)])) or None
        d = dq[dq["Opp Owner"] == o]
        won = d[d["__won"]]
        quota = get_target(tg, cql, o)
        commit = d[d["__open"] & (d["__bucket"] == "commit")]["Value"].sum()
        created = None
        if "Created Date" in d.columns:
            qstart = pd.Timestamp(dt.date(quarter_months(cfy, cq, sm)[0][0],
                                          quarter_months(cfy, cq, sm)[0][1], 1))
            created = df[(df["Opp Owner"] == o) & (df["Created Date"].notna())
                         & (df["Created Date"] >= qstart)]["Value"].sum()
        rows.append({
            "Sales Rep": o,
            "L2Q Attainment": (twon["Value"].sum() / tquota) if tquota else None,
            "L2Q Quota": tquota, "L2Q Bookings": twon["Value"].sum(),
            "L2Q ASP": (twon["Value"].sum() / len(twon)) if len(twon) else None,
            "CurQ Attainment": (won["Value"].sum() / quota) if quota else None,
            "CurQ Quota": quota, "CurQ Bookings": won["Value"].sum(),
            "CurQ Forecast": won["Value"].sum() + commit,
            "CurQ Weighted Pipe": d[d["__open"]]["Weighted"].sum(),
            "CurQ Total Pipe": d[d["__open"]]["Value"].sum(),
            "CurQ Pipe Created": created,
        })
    t = pd.DataFrame(rows)
    f = fmts_for(t.columns, cur)
    f.update({"L2Q Attainment": "0%", "CurQ Attainment": "0%"})
    for c in t.columns:
        if "Quota" in c or "Bookings" in c or "ASP" in c or "Pipe" in c or "Forecast" in c:
            f[c] = f'"{cur}"#,##0'
    row = write_df(ws, row, t, f)

    row = section_title(ws, row, "Open Pipeline — Owner × Stage", W)
    rows = []
    for o in owners:
        d = dfy[(dfy["Opp Owner"] == o) & dfy["__open"]]
        rec = {"Owner": o}
        for i, s in enumerate(cfg["stages"], start=1):
            sel = d[d["Stage"] == s["name"]]
            rec[f"S{i} {vl}"] = sel["Value"].sum()
            rec[f"S{i} Wtd"] = sel["Weighted"].sum()
            rec[f"S{i} #"] = len(sel)
        rows.append(rec)
    add_totals(rows)
    t = pd.DataFrame(rows)
    f = {c: f'"{cur}"#,##0' for c in t.columns if vl in c or "Wtd" in c}
    row = write_df(ws, row, t, f, total_row=True)

    if "Deal Band" in df.columns:
        row = section_title(ws, row, "Open Pipeline — Owner × Deal Band", W)
        bands = [b["name"] for b in sorted(cfg["deal_bands"], key=lambda b: b["min"])]
        rows = []
        for o in owners:
            d = dfy[(dfy["Opp Owner"] == o) & dfy["__open"]]
            rec = {"Owner": o}
            for b in bands:
                sel = d[d["Deal Band"] == b]
                rec[f"{b} {vl}"] = sel["Value"].sum()
                rec[f"{b} #"] = len(sel)
            rec[f"Total {vl}"] = d["Value"].sum()
            rec["Total #"] = len(d)
            rows.append(rec)
        add_totals(rows)
        t = pd.DataFrame(rows)
        f = {c: f'"{cur}"#,##0' for c in t.columns if vl in c}
        row = write_df(ws, row, t, f, total_row=True)

    row = section_title(ws, row, f"Deal Review — open opps FY{fy}", W)
    rev = dfy[dfy["__open"]].sort_values("Value", ascending=False)
    cols = [c for c in ["Opp Owner", "Forecast Category", "Stage", "Opportunity Name",
                        "Next Step", "Type", "Deal Band", "ICP Grade", "Close Date",
                        "Value", "Opp Link"] if c in rev.columns]
    t = rev[cols].copy()
    if len(t):
        t["Close Date"] = pd.to_datetime(t["Close Date"]).dt.strftime("%Y-%m-%d")
    write_df(ws, row, t, fmts_for(t.columns, cur))
    ws.freeze_panes = "A3"
    autosize(ws)


def tab_readme(wb, cfg, df, li, tg, asof, unknown):
    ws = wb.create_sheet("README", 0)
    W = 8
    row = section_title(ws, 1, f"{cfg.get('client_name', 'Client')} — GTM Dashboards", W)
    lines = [
        f"Built {dt.date.today():%d %b %Y}, as-of date {asof:%d %b %Y}.",
        f"Value measure: {cfg.get('value_label', 'Value')}. "
        f"Fiscal year starts month {cfg.get('fiscal_year_start_month', 1)}.",
        f"Source rows: {len(df)} opportunities"
        + (f", {len(li)} line items" if li is not None else ", no line items")
        + (", targets supplied." if tg is not None else ", NO TARGETS — attainment/coverage awaiting input."),
        "",
        "Tabs: 1. Quarterly Forecast · 2. FY Forecast & Pipeline · "
        "3. Seller Performance · raw data tabs at the end for traceability.",
        "",
        "ASSUMPTIONS TO CONFIRM WITH THE CLIENT:",
        "• Stage win probabilities (drive every Weighted number): "
        + ", ".join(f"{s['name']} {s['probability']:.0%}" for s in cfg["stages"]),
        "• Deal-band thresholds: "
        + ", ".join(f"{b['name']} ≥ {b['min']:,}" for b in
                    sorted(cfg.get("deal_bands", []), key=lambda b: b["min"])),
        "• Value field mapping (Amount vs custom ARR vs line-item rollup).",
    ]
    if unknown:
        lines.append(f"• DATA WARNING — stages in the data not in config "
                     f"(excluded from rollups): {', '.join(unknown)}")
    blanks = df[df["Opp Owner"].isna() | df["Stage"].isna()]
    if len(blanks):
        lines.append(f"• DATA WARNING — {len(blanks)} opportunities with blank "
                     f"owner or stage: {', '.join(map(str, blanks['Opp Id'].head(15)))}")
    r = row + 1
    for ln in lines:
        ws.cell(row=r, column=1, value=ln)
        r += 1
    ws.column_dimensions["A"].width = 110


def raw_tabs(wb, df, li, tg, cfg):
    for name, data in [("Raw Opportunities", df.drop(columns=[c for c in df.columns
                                                              if c.startswith("__")])),
                       ("Raw Line Items", li), ("Targets", tg)]:
        if data is None:
            continue
        ws = wb.create_sheet(name)
        out = data.copy()
        for c in out.columns:
            if pd.api.types.is_datetime64_any_dtype(out[c]):
                out[c] = out[c].dt.strftime("%Y-%m-%d")
        write_df(ws, 1, out)
        autosize(ws)
    ws = wb.create_sheet("Config")
    ws.cell(row=1, column=1, value=json.dumps(cfg, indent=2))
    ws.column_dimensions["A"].width = 80


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--opportunities", required=True)
    ap.add_argument("--line-items")
    ap.add_argument("--targets")
    ap.add_argument("--as-of", default=str(dt.date.today()))
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    asof = dt.date.fromisoformat(args.as_of)
    cfg, df, li, tg, unknown = load(args)

    wb = Workbook()
    wb.remove(wb.active)
    tab_quarterly(wb, cfg, df, tg, asof)
    tab_fy(wb, cfg, df, li, tg, asof)
    tab_sellers(wb, cfg, df, tg, asof)
    tab_readme(wb, cfg, df, li, tg, asof, unknown)
    raw_tabs(wb, df, li, tg, cfg)
    wb.save(args.output)

    checks = []
    fy = cfg.get("fiscal_year")
    won_fy = df[(df["__fy"] == fy) & df["__won"]]["Value"].sum()
    checks.append(f"FY bookings (recompute): {won_fy:,.0f}")
    checks.append(f"Open pipeline total: {df[df['__open']]['Value'].sum():,.0f}")
    if unknown:
        checks.append(f"WARNING unmapped stages: {unknown}")
    print(f"Saved {args.output}")
    print("\n".join(checks))


if __name__ == "__main__":
    main()
